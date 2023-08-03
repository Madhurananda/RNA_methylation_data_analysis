
import pandas as pd
import openpyxl
import os
from multiprocessing.pool import ThreadPool
from datetime import datetime
# import concurrent.futures
import time
import threading




# # '''
# # I have tried with the webcrawler, but didn't work. 
# # Now, I have copied it to a Excel file and trying to get the hyperlinks ... 
# # '''

# # def download_wget(args):
# def download_wget(url, save_dir):
#     t0 = time.time()
#     # url, save_dir = args[1], args[2]
#     # _, url, save_dir = args[0], args[1], args[2]
#     try:
#         wget_cmd = 'wget {} -P {}'.format(url, save_dir)
#         # wget_cmd ='wget https://sra-pub-src-1.s3.amazonaws.com/SRZ190746/IVT_m1A_fast5.tar.gz.1 -P /home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data_NEW/'
#         print('\n\nwget_cmd: ', wget_cmd)
#         # os.system(wget_cmd)
#         return(url, time.time() - t0)
#     except Exception as e:
#         print('Exception in download_url(): ', e)
    

# def download_parallel(args):
#     n_jobs = args[2][0]
#     # n_jobs, _, _ = args[0], args[1], args[2]
#     print('n_jobs: ', n_jobs)
#     results = ThreadPool(n_jobs).imap_unordered(download_wget, args)
#     for result in results:
#         print('url downloaded: ', result[0], 'time (s): ', result[1])



# if __name__=='__main__':

#     startTime = datetime.now()
#     current_time = startTime.strftime("%Y/%m/%d at %H:%M:%S")
#     print('\n\nThe script starting at: ' + str(current_time), ' \n\n' )
#     # startTime = time.time()

#     # basefolder = sys.argv[1]
#     # output_folder = sys.argv[2]
#     # zip_type = int(sys.argv[3])

    



#     xls_file = '/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data/Paper5_datasets.xlsx'
#     save_dir = '/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data_NEW/'
#     sh_file = '/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data/wget_fast5.sh'
#     n_jobs = 10

#     dataframe1 = pd.read_excel(xls_file)
    
#     # print(len(dataframe1))

#     wb = openpyxl.load_workbook(xls_file)
#     ws = wb['Sheet1']


#     file_list = []
#     f = open(sh_file, "w")
#     for a_row in range(len(dataframe1)):
#         # This will fail if there is no hyperlink to target
#         try:
#             hLink = ws.cell(row=(a_row+1), column=3).hyperlink.target
#             if 'FAST5' in hLink.upper():
#                 ## Add another check depending on the dataset: 
#                 if not 'CDNA' in hLink.upper():
#                     # print('The link is: ', hLink)
#                     f.write("wget "+hLink+" &" +"\n" )
#                     file_list.append(hLink)
#         except:
#             hLink = ws.cell(row=(a_row+1), column=3).value

#     f.write("wait\n")
#     f.close()


#     # inputs = zip(file_list, [save_dir]*len(file_list), [n_jobs]*len(file_list))
#     inputs = (file_list, [save_dir]*len(file_list), [n_jobs]*len(file_list))

#     # download_parallel(inputs)

#     ## Trying manual Threads (working fine)
#     threads = []
#     for _ in range(20):
#         t = threading.Thread(target=download_wget, args=(file_list, [save_dir]*len(file_list)))
#         t.start()
#         threads.append(t)

#     for thread in threads:
#         thread.join()


#     executionTime = (datetime.now() - startTime)
    
#     # now = datetime.now()
#     current_time = datetime.now().strftime("%Y/%m/%d at %H:%M:%S")
#     print('\n\nThe script completed at: ' + str(current_time))
#     print('Execution time: ' + str(executionTime), ' \n\n')



import requests
import time
from multiprocessing import cpu_count
from multiprocessing.pool import ThreadPool, Pool
import os


def download_url(args):
    t0 = time.time()
    url, save_dir = args[0], args[1]
    # url = list(zip(*args))[0]
    # save_dir = list(zip(*args))[1]
    try:
        wget_cmd = 'wget {} -P {}'.format(url, save_dir)
#         # wget_cmd ='wget https://sra-pub-src-1.s3.amazonaws.com/SRZ190746/IVT_m1A_fast5.tar.gz.1 -P /home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data_NEW/'
        print('\n\nwget_cmd: ', wget_cmd)
        os.system(wget_cmd)
    #     r = requests.get(url)
    #     with open(fn, 'wb') as f:
    #         f.write(r.content)
        return(url, time.time() - t0)
    except Exception as e:
        print('Exception in download_url():', e)

# def download_parallel_ThreadPool(args):
#     # cpus = cpu_count()
#     n_jobs = list(zip(*args))[2][0]
#     print('n_jobs: ', n_jobs)
#     results = ThreadPool(n_jobs).imap_unordered(download_url, args)
#     for result in results:
#         print('url:', result[0], 'time (s):', result[1])

def download_parallel_processPool(args):
    n_jobs = 20
    # n_jobs = list(zip(*args))[2][0]
    # n_jobs = args[2]
    print('n_jobs: ', n_jobs)
    results = Pool(n_jobs).imap_unordered(download_url, args)
    for result in results:
        print('url:', result[0], 'time (s):', result[1])

# urls = ['https://www.northwestknowledge.net/metdata/data/pr_1979.nc',
# 'https://www.northwestknowledge.net/metdata/data/pr_1980.nc',
# 'https://www.northwestknowledge.net/metdata/data/pr_1981.nc',
# 'https://www.northwestknowledge.net/metdata/data/pr_1982.nc']

# fns = [r'/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data_NEW/pr_1979.nc',
# r'/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data_NEW/pr_1980.nc',
# r'/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data_NEW/pr_1981.nc',
# r'/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data_NEW/pr_1982.nc']

# inputs = zip(urls, fns)



if __name__=='__main__':

    xls_file = '/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data/Paper5_datasets.xlsx'
    save_dir = '/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data_NEW/'
    sh_file = '/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data/wget_fast5.sh'
    # n_jobs = 10

    dataframe1 = pd.read_excel(xls_file)
    
    # print(len(dataframe1))

    wb = openpyxl.load_workbook(xls_file)
    ws = wb['Sheet1']


    file_list = []
    # save_dir_list = [] 
    # f = open(sh_file, "w")
    for a_row in range(len(dataframe1)):
        # This will fail if there is no hyperlink to target
        try:
            hLink = ws.cell(row=(a_row+1), column=3).hyperlink.target
            if 'FAST5' in hLink.upper():
                ## Add another check depending on the dataset: 
                if not 'CDNA' in hLink.upper():
                    # print('The link is: ', hLink)
                    # f.write("wget "+hLink+" &" +"\n" )
                    file_list.append(hLink)
        except:
            hLink = ws.cell(row=(a_row+1), column=3).value

    # f.write("wait\n")
    # f.close()
    # print(file_list)

    file_list = ['https://sra-pub-src-1.s3.amazonaws.com/SRZ190740/HS_rRNA_dRNA_fast5.tar.gz.1',
    'https://sra-pub-src-1.s3.amazonaws.com/SRZ190743/IVT_hm5C_fast5.tar.gz.1',
    'https://sra-pub-src-1.s3.amazonaws.com/SRZ190746/IVT_m1A_fast5.tar.gz.1']

    inputs = zip(file_list, [save_dir]*len(file_list))
    # inputs = zip(file_list, [save_dir]*len(file_list), [n_jobs]*len(file_list))




# t0 = time.time()
# for i in inputs:
#     result = download_url(i)
#     print('url:', result[0], 'time:', result[1])
# print('Total time:', time.time() - t0)

# download_parallel_ThreadPool(inputs)


download_parallel_processPool(inputs)
