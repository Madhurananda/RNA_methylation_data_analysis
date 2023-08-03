
import pandas as pd
import openpyxl
import os
from multiprocessing.pool import ThreadPool, Pool
from datetime import datetime
from multiprocessing import cpu_count
# import time
import requests
import urllib.request
# import wget
from tqdm import tqdm
import fnmatch
from ftplib import FTP

import sys
sys.path.insert(0, '/home/madhu/datasets/codes')
from main import check_endWith, check_create_dir, check_conda_env

'''
I am using multiple CPU to download a large number of files ... 
It uses maximum number of CPU possible. 
'''


def download_url(args):
    t0 = datetime.now()
    url, save_dir = args[0], args[1]
    # url = list(zip(*args))[0]
    # save_dir = list(zip(*args))[1]
    exist = False
    file = save_dir+url.split('/')[-1]
    # print(file)
    try:
        if os.path.isfile(file):
            print('\n\nThe file: ', file, ' already exists ...')
            exist = True
        else:
            wget_cmd = 'wget {} -P {}'.format(url, save_dir)
    # #         # wget_cmd ='wget https://sra-pub-src-1.s3.amazonaws.com/SRZ190746/IVT_m1A_fast5.tar.gz.1 -P /home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data_NEW/'
            print('\n\nwget_cmd: ', wget_cmd)
            os.system(wget_cmd)
    
    except Exception as e:
        print('Exception in download_url():', e)
    pbar.update(1)
    return(url, datetime.now() - t0, exist)

# def download_url_python(args):
#     t0 = datetime.now()
#     url, save_dir = args[0], args[1]
#     try:
#         ftp = FTP('ftp://ftp.sra.ebi.ac.uk')

#         #if no username and password is required.
#         ftp.login()

#         # if you have to change directory on FTP server.
#         ftp.cwd('/vol1/run/ERR376/ERR3764345/')

#         # Get all files
#         files = ftp.nlst()
#         print('files: ', files)
#         # Download files
#         for file in files:
#             if fnmatch.fnmatch(file, '*.tar.gz'):   #To download specific files.
#                 print("Downloading..." + file)

#                 try:
#                     ftp.retrbinary("RETR " + file ,open(save_dir + file, 'wb').write)

#                 except EOFError:    # To avoid EOF errors.
#                     pass

#         ftp.close()

#         # urllib.request.urlretrieve(url, 'test')

#         # wget.download(url, out=save_dir)

#         # r = requests.get(url, allow_redirects=True)
#         # open(save_dir+'test.tar.gz', 'wb').write(r.content)
#         # print('\n\nURL: ', url, ' saved at: ', save_dir)

#         # wget_cmd = 'wget {} -P {}'.format(url, save_dir)
#         # print('\n\nwget_cmd: ', wget_cmd)
#         # os.system(wget_cmd)
        
    # except Exception as e:
    #     print('Exception in download_url():', e)

    # return(url, datetime.now() - t0)






if __name__=='__main__':

    startTime = datetime.now()
    current_time = startTime.strftime("%Y/%m/%d at %H:%M:%S")
    print('\n\nThe script starting at: ' + str(current_time), ' \n\n' )

    # xls_file = '/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data/Paper5_datasets.xlsx'
    # save_dir = '/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data_NEW/'
    # # sh_file = '/home/madhu/datasets/download/p5_decoding_epitranscriptional_native_RNA_seq/original_data/wget_fast5.sh'
    # n_jobs = 20

    save_dir = sys.argv[1]
    # xls_file = sys.argv[2]
    
    save_dir = check_endWith(save_dir)

    # dataframe1 = pd.read_excel(xls_file)
    # # print(len(dataframe1))
    # wb = openpyxl.load_workbook(xls_file)
    # ws = wb['Sheet1']
    # file_list = []
    # for a_row in range(len(dataframe1)):
    #     # This will fail if there is no hyperlink to target
    #     try:
    #         hLink = ws.cell(row=(a_row+1), column=3).hyperlink.target
    #         if 'FAST5' in hLink.upper():
    #             ## Add another check depending on the dataset: 
    #             if not 'CDNA' in hLink.upper():
    #                 # print('The link is: ', hLink)
    #                 file_list.append(hLink)
    #     except:
    #         hLink = ws.cell(row=(a_row+1), column=3).value

    

    # file_list = ['https://sra-pub-src-1.s3.amazonaws.com/SRZ190740/HS_rRNA_dRNA_fast5.tar.gz.1',
    # 'https://sra-pub-src-1.s3.amazonaws.com/SRZ190743/IVT_hm5C_fast5.tar.gz.1',
    # 'https://sra-pub-src-1.s3.amazonaws.com/SRZ190746/IVT_m1A_fast5.tar.gz.1']

    file_list = ['ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764345/col0_nanopore_drs_1.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764346/col0_nanopore_drs_with_adapter_1.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764347/col0_nanopore_drs_2a.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764348/col0_nanopore_drs_2b.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764349/col0_nanopore_drs_3.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764350/col0_nanopore_drs_with_adapter_3.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764351/col0_nanopore_drs_4.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764352/vir1_nanopore_drs_1.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764353/vir1_nanopore_drs_2.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764354/vir1_nanopore_drs_3.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764355/vir1_nanopore_drs_4.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764356/VIRc_nanopore_drs_1.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764357/VIRc_nanopore_drs_2.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764358/VIRc_nanopore_drs_3.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764359/VIRc_nanopore_drs_4.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR377/ERR3772299/IVT_GFP.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR377/ERR3772300/MALAT1_drs_cntrl.tar.gz', 'ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR377/ERR3772301/MALAT1_drs_m6a.tar.gz']
    
    # file_list = ['ftp://ftp.sra.ebi.ac.uk/vol1/run/ERR376/ERR3764345/col0_nanopore_drs_1.tar.gz']

    inputs = zip(file_list, [save_dir]*len(file_list))
    # print(len(file_list))
    pbar = tqdm(total=len(file_list))

    ## Use the maximum CPU possible 
    cpus = cpu_count()
    n_jobs = min(cpus, len(file_list))
    print('n_jobs: ', n_jobs)
    results = Pool(n_jobs).imap_unordered(download_url, inputs)
    # results = ThreadPool(n_jobs).imap_unordered(download_url_python, inputs)
    for result in results:
        # print('url:', result[0], 'time (s):', result[1], '  ... Exist: ', result[2])
        # print(result[2])
        if result[2] == False:
            print('url:', result[0], 'time (s):', result[1], '  ... Exist: ', result[2])
            # print('This already exisits: ', result[0])

    executionTime = (datetime.now() - startTime)
    current_time = datetime.now().strftime("%Y/%m/%d at %H:%M:%S")
    print('\n\nThe script completed at: ' + str(current_time))
    print('Execution time: ' + str(executionTime), ' \n\n')
